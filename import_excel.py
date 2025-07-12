import os
import pandas as pd
import sqlite3
import openpyxl
from openpyxl.utils import get_column_letter

excel_file = 'output.xlsx'
output_dir = 'static/images'
os.makedirs(output_dir, exist_ok=True)

# 读取Excel数据
df = pd.read_excel(excel_file)

# 规范列名（中文转英文，方便后续处理）
df.rename(columns={
    '标题': 'title',
    '价格': 'price',
    'QQ': 'qq',
    '微信': 'wechat',
    '手机': 'phone',
    '帖子链接': 'post_link'
}, inplace=True)

# 加载openpyxl，读取嵌入图片
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 映射：行号 -> 图片列表 (cell, image_obj)
image_map = {}
for image in ws._images:
    if not hasattr(image, 'anchor') or not hasattr(image.anchor, '_from'):
        continue
    anchor = image.anchor._from
    row = anchor.row + 1  # 行号，从1开始
    cell = f"{get_column_letter(anchor.col + 1)}{row}"
    image_map.setdefault(row, []).append((cell, image))

# 新增图片字段，初始化为空
df['image1'] = ''
df['image2'] = ''
df['image3'] = ''

for idx, row in df.iterrows():
    title = str(row['title']).strip()
    if not title:
        title = f"row_{idx+2}"
    # 这里改为“标题_行号”作为文件夹名，避免重复文件夹
    safe_title = title.replace('/', '_').replace('\\', '_').replace(' ', '_')
    folder_name = f"{safe_title}_{idx+2}"
    folder_path = os.path.join(output_dir, folder_name)
    os.makedirs(folder_path, exist_ok=True)

    excel_row_num = idx + 2  # Excel表头第1行，数据从第2行开始
    imgs = image_map.get(excel_row_num, [])

    # 保存图片
    for i, (cell, img) in enumerate(imgs[:4]):
        img_path = os.path.join(folder_path, f'F{i+1}.png')
        with open(img_path, 'wb') as f:
            f.write(img._data())
        df.at[idx, f'image{i+1}'] = f'/static/images/{folder_name}/F{i+1}.png'

# 转换价格列为数字，无法转换设为None
df['price'] = pd.to_numeric(df['price'], errors='coerce')

# 写入SQLite数据库
conn = sqlite3.connect('data.db')
df.to_sql('data', conn, if_exists='replace', index=False)
conn.close()

print("✅ 数据和图片导入完成！文件夹名带行号避免重复")
