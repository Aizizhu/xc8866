import argparse
import os
import random
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Referer": "https://xc8866.com/",
    "Accept-Language": "zh-CN,zh;q=0.9",
}


@dataclass(slots=True)
class CrawlConfig:
    start_url: str
    total_pages: int
    threads: int = 6
    output_xlsx: str = "output.xlsx"
    image_dir: str = "images"
    crawled_file: str = "crawled_posts.txt"
    min_delay: float = 0.8
    max_delay: float = 1.5
    request_timeout: tuple[int, int] = (3, 6)
    flush_batch: int = 10


@dataclass(slots=True)
class PostRecord:
    title: str
    price: str
    qq: str
    wechat: str
    phone: str
    post_url: str
    image_files: list[str]


class XC8866Crawler:
    def __init__(self, config: CrawlConfig) -> None:
        self.config = config
        self.session = requests.Session()
        self.session.headers.update(DEFAULT_HEADERS)

        self.output_path = Path(config.output_xlsx)
        self.image_root = Path(config.image_dir)
        self.crawled_path = Path(config.crawled_file)

        self.image_root.mkdir(parents=True, exist_ok=True)

        self.excel_lock = threading.Lock()
        self.crawled_lock = threading.Lock()

    @staticmethod
    def log(msg: str) -> None:
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}")

    @staticmethod
    def sanitize_filename(name: str) -> str:
        return re.sub(r"[\\/:*?\"<>|]", "_", name)

    def load_crawled(self) -> set[str]:
        if not self.crawled_path.exists():
            return set()

        crawled: set[str] = set()
        with self.crawled_path.open("r", encoding="utf-8") as file:
            for line in file:
                line = line.strip()
                if not line:
                    continue
                crawled.add(line.split("\t")[0])
        return crawled

    def save_crawled(self, post_id: str, post_url: str) -> None:
        with self.crawled_lock:
            with self.crawled_path.open("a", encoding="utf-8") as file:
                file.write(f"{post_id}\t{post_url}\n")

    @staticmethod
    def normalize_url(url: str, base_url: str) -> str:
        if url.startswith("//"):
            return "https:" + url
        if url.startswith("/"):
            return urljoin(base_url, url)
        return url

    @staticmethod
    def extract_contact_by_regex(text: str, patterns: list[str]) -> str:
        for pattern in patterns:
            match = re.search(pattern, text, flags=re.IGNORECASE)
            if match:
                return match.group(1).strip()
        return ""

    def extract_info_from_table(self, soup: BeautifulSoup) -> tuple[str, str, str, str]:
        price = qq = wechat = phone = ""

        def set_field(label: str, value: str) -> None:
            nonlocal price, qq, wechat, phone
            if not value:
                return
            if "价格" in label and not price:
                price = value
            elif "QQ" in label.upper() and not qq:
                qq = value
            elif "微信" in label and not wechat:
                wechat = value
            elif ("电话" in label or "手机" in label) and not phone:
                phone = value

        for row in soup.select("table tr"):
            label = ""
            label_el = row.find(["th", "td"])
            value_el = label_el.find_next_sibling(["th", "td"]) if label_el else None
            if label_el and value_el:
                label = label_el.get_text(" ", strip=True)
                value = value_el.get_text(" ", strip=True)
                set_field(label, value)

        for item in soup.select("dl dt, li, div"):
            line = item.get_text(" ", strip=True)
            if "：" not in line and ":" not in line:
                continue
            parts = re.split(r"[：:]", line, maxsplit=1)
            if len(parts) != 2:
                continue
            set_field(parts[0], parts[1])

        full_text = soup.get_text("\n", strip=True)
        if not price:
            price = self.extract_contact_by_regex(full_text, [r"价格\s*[：:]\s*([^\n\r]+)"])
        if not qq:
            qq = self.extract_contact_by_regex(full_text, [r"QQ\s*[：:]\s*([0-9A-Za-z_-]{5,20})"])
        if not wechat:
            wechat = self.extract_contact_by_regex(full_text, [r"微信\s*[：:]\s*([0-9A-Za-z_-]{5,40})"])
        if not phone:
            phone = self.extract_contact_by_regex(full_text, [r"(?:电话|手机)\s*[：:]\s*([0-9+\-\s]{7,20})"])

        return price, qq, wechat, phone

    def extract_images(self, soup: BeautifulSoup, page_url: str) -> list[str]:
        image_urls: list[str] = []
        seen: set[str] = set()
        valid_exts = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}
        src_keys = ("src", "data-src", "data-original", "data-echo", "data-lazy-src")

        for img in soup.find_all("img"):
            src = ""
            for key in src_keys:
                value = img.get(key, "").strip()
                if value:
                    src = value
                    break

            if not src:
                srcset = img.get("srcset", "").strip()
                if srcset:
                    src = srcset.split(",")[0].strip().split(" ")[0]

            if not src:
                continue

            src_lower = src.lower()
            if any(x in src_lower for x in ("zwzp.jpg", "default.jpg", "nopic.jpg", "avatar", "logo", "icon")):
                continue

            img_url = self.normalize_url(src, page_url)
            ext = os.path.splitext(urlparse(img_url).path)[-1].lower()
            if ext and ext not in valid_exts:
                continue

            if img_url not in seen:
                seen.add(img_url)
                image_urls.append(img_url)

        return image_urls

    def parse_post(self, post_url: str) -> PostRecord | None:
        try:
            response = self.session.get(post_url, timeout=self.config.request_timeout)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, "html.parser")

            meta_title = soup.find("meta", attrs={"property": "og:title"})
            if meta_title and meta_title.get("content"):
                title = meta_title["content"].strip()
            else:
                meta_desc = soup.find("meta", attrs={"name": "description"})
                if meta_desc and meta_desc.get("content"):
                    title = meta_desc["content"].strip()
                else:
                    title_tag = soup.select_one("h1, h2, h3, h4.break-all, .thread-title, .topic-title")
                    title = title_tag.get_text(" ", strip=True) if title_tag else "标题未找到"

            price, qq, wechat, phone = self.extract_info_from_table(soup)
            image_urls = self.extract_images(soup, post_url)

            return PostRecord(
                title=title,
                price=price,
                qq=qq,
                wechat=wechat,
                phone=phone,
                post_url=post_url,
                image_files=self.download_images(image_urls, self.build_post_image_dir(post_url)),
            )
        except Exception as exc:  # noqa: BLE001
            self.log(f"访问帖子失败: {post_url} 错误: {exc}")
            return None

    def build_post_image_dir(self, post_url: str) -> Path:
        post_id = post_url.rstrip("/").split("/")[-1].replace(".htm", "")
        safe_post_id = self.sanitize_filename(post_id)
        image_dir = self.image_root / safe_post_id
        image_dir.mkdir(parents=True, exist_ok=True)
        return image_dir

    def download_images(self, image_urls: Iterable[str], image_dir: Path) -> list[str]:
        downloaded_files: list[str] = []

        for index, img_url in enumerate(image_urls, start=1):
            ext = os.path.splitext(urlparse(img_url).path)[-1].lower()
            if not re.match(r"\.(jpg|jpeg|png|gif|bmp|webp)$", ext):
                ext = ".jpg"

            image_name = f"{index}{ext}"
            image_path = image_dir / image_name

            if image_path.exists():
                downloaded_files.append(str(image_path))
                self.log(f"  跳过已存在图片: {image_name}")
                continue

            try:
                response = self.session.get(
                    img_url,
                    timeout=self.config.request_timeout,
                    stream=True,
                )
                response.raise_for_status()
                with image_path.open("wb") as file:
                    for chunk in response.iter_content(chunk_size=1024):
                        if chunk:
                            file.write(chunk)
                downloaded_files.append(str(image_path))
                self.log(f"  下载图片: {image_name}")
                time.sleep(random.uniform(0.2, 0.4))
            except Exception as exc:  # noqa: BLE001
                self.log(f"图片下载失败: {img_url}, 错误: {exc}")

        return downloaded_files

    def append_records_to_excel(self, records: list[PostRecord]) -> None:
        if not records:
            return

        max_imgs = max(3, max(len(record.image_files) for record in records))
        headers = ["标题", "价格", "QQ", "微信", "手机"] + [f"图片{i}" for i in range(1, max_imgs + 1)] + ["帖子链接"]

        if self.output_path.exists():
            workbook = load_workbook(self.output_path)
            worksheet = workbook.active
            existing_headers = [cell.value for cell in worksheet[1]]
            if len(existing_headers) < len(headers):
                for col_idx in range(len(existing_headers) + 1, len(headers) + 1):
                    worksheet.cell(row=1, column=col_idx, value=headers[col_idx - 1])
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "爬取结果"
            worksheet.append(headers)

        for record in records:
            row_values = [record.title, record.price, record.qq, record.wechat, record.phone]
            worksheet.append(row_values + [""] * max_imgs + [record.post_url])
            row_idx = worksheet.max_row

            for i, image_path in enumerate(record.image_files[:max_imgs]):
                try:
                    PILImage.open(image_path).verify()
                    excel_image = XLImage(image_path)
                    excel_image.width = 100
                    excel_image.height = 100
                    col_letter = chr(ord("F") + i)
                    worksheet.add_image(excel_image, f"{col_letter}{row_idx}")
                except Exception as exc:  # noqa: BLE001
                    self.log(f"❌ 图片插入失败: {image_path}, 错误: {exc}")

        workbook.save(self.output_path)
        self.log(f"✅ 写入 Excel：{self.output_path}")

    @staticmethod
    def get_page_threads(soup: BeautifulSoup) -> list[str]:
        links: list[str] = []
        seen: set[str] = set()

        for thread in soup.select("li.media.thread.tap[data-href], li[data-href], [data-href]"):
            href = thread.get("data-href", "").strip()
            if href and href not in seen:
                seen.add(href)
                links.append(href)

        selectors = [
            'a[href*="/thread-"]',
            'a[href*="/topics/"]',
            'a[href*="/topic/"]',
            'a[href$=".htm"]',
        ]
        for selector in selectors:
            for anchor in soup.select(selector):
                href = anchor.get("href", "").strip()
                if not href:
                    continue
                if href.startswith("javascript:") or href.startswith("#"):
                    continue
                if href not in seen:
                    seen.add(href)
                    links.append(href)

        return links

    @staticmethod
    def is_post_link(link: str) -> bool:
        lower_link = link.lower()
        return any(token in lower_link for token in ("/thread-", "/topic/", "/topics/")) or lower_link.endswith(".htm")

    def crawl_single_page(self, page_url: str, page_num: int, crawled_posts: set[str]) -> None:
        self.log(f"📄 线程爬取第 {page_num} 页：{page_url}")
        batch: list[PostRecord] = []

        try:
            response = self.session.get(page_url, timeout=self.config.request_timeout)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, "html.parser")

            links = [link for link in self.get_page_threads(soup) if self.is_post_link(link)]
            if not links:
                self.log(f"⚠️ 第 {page_num} 页没有获取到帖子链接，跳过")
                return

            self.log(f"🔍 本页共发现 {len(links)} 条帖子链接")
            for idx, link in enumerate(links, start=1):
                post_id = link.replace(".htm", "").replace("/", "_")
                if post_id in crawled_posts:
                    self.log(f"跳过已爬取帖子 {post_id} ({link})")
                    continue

                post_url = urljoin("https://xc8866.com/", link)
                self.log(f"➡️ 正在爬取帖子 {idx}/{len(links)}: {post_url}")
                record = self.parse_post(post_url)
                if not record:
                    self.log(f"⚠️ 帖子解析失败，跳过: {post_url}")
                    continue

                self.log(f"  标题: {record.title}")
                self.log(f"  下载图片 {len(record.image_files)} 张")

                batch.append(record)
                self.save_crawled(post_id, post_url)
                crawled_posts.add(post_id)

                if len(batch) >= self.config.flush_batch:
                    with self.excel_lock:
                        self.append_records_to_excel(batch)
                    self.log(f"✅ 已保存 {len(batch)} 条帖子数据")
                    batch.clear()

                time.sleep(random.uniform(self.config.min_delay, self.config.max_delay))

            if batch:
                with self.excel_lock:
                    self.append_records_to_excel(batch)
                self.log(f"✅ 本页剩余 {len(batch)} 条帖子数据已保存")

        except Exception as exc:  # noqa: BLE001
            self.log(f"爬取页面失败: {page_url} 错误: {exc}")

    @staticmethod
    def build_page_urls(start_url: str, total_pages: int) -> list[tuple[int, str]]:
        parsed = urlparse(start_url)
        query_pairs = parse_qsl(parsed.query, keep_blank_values=True)
        query_map = dict(query_pairs)

        if "page" in query_map:
            try:
                start_page = int(query_map["page"])
            except ValueError as exc:
                raise ValueError("起始链接中的 page 参数必须是数字") from exc

            urls: list[tuple[int, str]] = []
            for page_num in range(start_page, start_page + total_pages):
                current_pairs = [
                    (key, str(page_num) if key == "page" else value)
                    for key, value in query_pairs
                ]
                rebuilt = parsed._replace(query=urlencode(current_pairs))
                urls.append((page_num, urlunparse(rebuilt)))
            return urls

        match = re.search(r"forum-23-(\d+)\.htm", start_url)
        if match:
            start_page = int(match.group(1))
            urls = []
            for page_num in range(start_page, start_page + total_pages):
                url = re.sub(r"forum-23-\d+\.htm", f"forum-23-{page_num}.htm", start_url)
                urls.append((page_num, url))
            return urls

        raise ValueError("起始链接格式不正确，应包含 page 参数（如 ?page=1）")

    def crawl(self) -> None:
        crawled_posts = self.load_crawled()
        page_urls = self.build_page_urls(self.config.start_url, self.config.total_pages)

        with ThreadPoolExecutor(max_workers=self.config.threads) as executor:
            future_map = {
                executor.submit(self.crawl_single_page, url, page_num, crawled_posts): page_num
                for page_num, url in page_urls
            }
            for future in as_completed(future_map):
                page_num = future_map[future]
                try:
                    future.result()
                    self.log(f"✅ 第 {page_num} 页爬取完成")
                except Exception as exc:  # noqa: BLE001
                    self.log(f"❌ 第 {page_num} 页爬取异常: {exc}")

        self.log("✅ 所有任务完成，程序退出")


def parse_args() -> CrawlConfig:
    parser = argparse.ArgumentParser(description="xc8866 爬虫：抓取帖子、图片并写入 Excel")
    parser.add_argument("--start-url", type=str, required=True, help="起始页链接")
    parser.add_argument("--total-pages", type=int, required=True, help="总共需要爬取多少页")
    parser.add_argument("--threads", type=int, default=6, help="最大线程数，默认 6")
    parser.add_argument("--output", type=str, default="output.xlsx", help="Excel 输出文件，默认 output.xlsx")
    parser.add_argument("--images-dir", type=str, default="images", help="图片输出目录，默认 images")
    parser.add_argument("--state-file", type=str, default="crawled_posts.txt", help="断点状态文件，默认 crawled_posts.txt")
    parser.add_argument("--flush-batch", type=int, default=10, help="累计多少条写入一次 Excel，默认 10")

    args = parser.parse_args()

    return CrawlConfig(
        start_url=args.start_url,
        total_pages=args.total_pages,
        threads=max(1, args.threads),
        output_xlsx=args.output,
        image_dir=args.images_dir,
        crawled_file=args.state_file,
        flush_batch=max(1, args.flush_batch),
    )


def main() -> None:
    config = parse_args()
    crawler = XC8866Crawler(config)
    crawler.crawl()


if __name__ == "__main__":
    main()
